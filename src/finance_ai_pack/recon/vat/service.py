from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from finance_ai_pack.config import Settings
from finance_ai_pack.connectors.odoo.client import OdooClient
from finance_ai_pack.connectors.odoo.fixtures_adapter import FixturesAdapter
from finance_ai_pack.connectors.odoo.live_adapter import LiveOdooAdapter


@dataclass
class TraMonthlyRow:
    period: str
    input_vat: float
    output_vat: float


def _iter_periods(period_from: str, period_to: str) -> list[str]:
    start = datetime.strptime(f"{period_from}-01", "%Y-%m-%d")
    end = datetime.strptime(f"{period_to}-01", "%Y-%m-%d")
    if start > end:
        raise ValueError("period_from must be <= period_to")

    periods = []
    cursor = start
    while cursor <= end:
        periods.append(cursor.strftime("%Y-%m"))
        if cursor.month == 12:
            cursor = datetime(cursor.year + 1, 1, 1)
        else:
            cursor = datetime(cursor.year, cursor.month + 1, 1)
    return periods


def _build_adapter(settings: Settings, fixtures_dir: Path):
    if settings.fixture_mode:
        return FixturesAdapter(fixtures_dir)
    return LiveOdooAdapter(OdooClient(settings))


def _validate_tra_columns(columns: set[str]) -> None:
    required = {"period", "input_vat", "output_vat"}
    if not required.issubset(columns):
        raise ValueError("TRA file must include columns: period,input_vat,output_vat")


def _parse_tra_csv(tra_file: Path) -> dict[str, TraMonthlyRow]:
    monthly: dict[str, TraMonthlyRow] = {}
    with tra_file.open(newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = {name.strip() for name in (reader.fieldnames or [])}
        _validate_tra_columns(fieldnames)
        for row in reader:
            period = (row.get("period") or "").strip()
            if not period:
                continue
            monthly[period] = TraMonthlyRow(
                period=period,
                input_vat=float(row.get("input_vat", 0) or 0),
                output_vat=float(row.get("output_vat", 0) or 0),
            )
    return monthly


def _parse_tra_xlsx(tra_file: Path) -> dict[str, TraMonthlyRow]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Excel TRA import requires openpyxl. Install openpyxl or provide CSV.") from exc

    wb = load_workbook(tra_file, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        return {}

    headers = [str(x).strip() if x is not None else "" for x in header_row]
    _validate_tra_columns(set(headers))
    index = {name: idx for idx, name in enumerate(headers)}

    monthly: dict[str, TraMonthlyRow] = {}
    for values in rows:
        if values is None:
            continue
        period_value = values[index["period"]] if index["period"] < len(values) else None
        period = str(period_value).strip() if period_value is not None else ""
        if not period:
            continue
        input_value = values[index["input_vat"]] if index["input_vat"] < len(values) else 0
        output_value = values[index["output_vat"]] if index["output_vat"] < len(values) else 0
        monthly[period] = TraMonthlyRow(
            period=period,
            input_vat=float(input_value or 0),
            output_vat=float(output_value or 0),
        )
    return monthly


def read_tra_file(tra_file: Path) -> dict[str, TraMonthlyRow]:
    suffix = tra_file.suffix.lower()
    if suffix == ".csv":
        return _parse_tra_csv(tra_file)
    if suffix == ".xlsx":
        return _parse_tra_xlsx(tra_file)
    raise ValueError("TRA file must be .csv or .xlsx")


def _categorize_exception(item: dict, period: str) -> str | None:
    explicit = (item.get("exception_hint") or "").lower()
    move_type = (item.get("move_type") or "").lower()
    source_period = item.get("source_period")
    has_ref = bool(item.get("document_ref"))

    if "credit" in explicit or move_type in {"out_refund", "in_refund"}:
        return "credit_notes/reversals"
    if "timing" in explicit or (source_period and source_period != period):
        return "timing/posting period"
    if not has_ref or "missing" in explicit:
        return "missing documents"
    if "tag" in explicit:
        return "wrong tax tags"
    if "fx" in explicit:
        return "FX rounding"
    return None


def reconcile_vat(
    period_from: str,
    period_to: str,
    fixtures_dir: Path,
    settings: Settings | None = None,
    tra_file: Path | None = None,
) -> dict:
    settings = settings or Settings.from_env()
    adapter = _build_adapter(settings, fixtures_dir)

    periods = _iter_periods(period_from, period_to)
    if not tra_file:
        default_csv = fixtures_dir / "vat" / f"tra_vat_{period_from}.csv"
        default_xlsx = fixtures_dir / "vat" / f"tra_vat_{period_from}.xlsx"
        if default_csv.exists():
            tra_file = default_csv
        elif default_xlsx.exists():
            tra_file = default_xlsx

    tra_by_month = read_tra_file(tra_file) if tra_file else {}

    monthly_summary = []
    exceptions = []
    net_diff_abs_total = 0.0

    for period in periods:
        input_lines = adapter.get_vat_tax_lines(period=period, vat_type="input")
        output_lines = adapter.get_vat_tax_lines(period=period, vat_type="output")

        odoo_input = round(sum(float(x.get("vat_amount", 0.0)) for x in input_lines), 2)
        odoo_output = round(sum(float(x.get("vat_amount", 0.0)) for x in output_lines), 2)

        tra_row = tra_by_month.get(period, TraMonthlyRow(period=period, input_vat=0.0, output_vat=0.0))
        tra_input = round(float(tra_row.input_vat), 2)
        tra_output = round(float(tra_row.output_vat), 2)

        input_diff = round(odoo_input - tra_input, 2)
        output_diff = round(odoo_output - tra_output, 2)
        net_diff = round((odoo_output - odoo_input) - (tra_output - tra_input), 2)
        net_diff_abs_total += abs(net_diff)

        control = adapter.get_vat_control_balance(period)
        monthly_summary.append(
            {
                "period": period,
                "odoo_input_vat": odoo_input,
                "tra_input_vat": tra_input,
                "input_difference": input_diff,
                "odoo_output_vat": odoo_output,
                "tra_output_vat": tra_output,
                "output_difference": output_diff,
                "net_vat_difference": net_diff,
                "vat_control_balance": round(float(control.get("closing_balance", 0.0)), 2),
                "vat_control_assumption": control.get("assumption", "best-effort"),
            }
        )

        for item in [*input_lines, *output_lines]:
            category = _categorize_exception(item, period=period)
            if not category:
                continue
            exceptions.append(
                {
                    "period": period,
                    "category": category,
                    "document_ref": item.get("document_ref", ""),
                    "source_period": item.get("source_period", ""),
                    "vat_amount": round(float(item.get("vat_amount", 0.0)), 2),
                    "tax_type": item.get("tax_type", ""),
                    "notes": item.get("notes", ""),
                }
            )

    narrative = (
        "Draft-only VAT reconciliation generated from Odoo extraction and TRA import. "
        "No auto-posting performed; numbers are deterministic from source records."
    )

    return {
        "command": "vat_pack",
        "period_from": period_from,
        "period_to": period_to,
        "mode": "fixture-only" if settings.fixture_mode else "live-odoo",
        "auto_posting": False,
        "narrative": narrative,
        "monthly_summary": monthly_summary,
        "exception_register": exceptions,
        "tra_file": str(tra_file) if tra_file else None,
        "metrics": {
            "months": len(monthly_summary),
            "exception_count": len(exceptions),
            "aggregate_net_vat_difference_abs": round(net_diff_abs_total, 2),
        },
    }
